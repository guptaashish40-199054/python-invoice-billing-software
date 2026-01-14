import tkinter as tk
from tkinter import messagebox
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# ---------------- CONFIG ----------------
VENDOR_NAME = "TECHCOM INFOSYS"
VENDOR_ADDR = "Delhi, India"
VENDOR_PHONE = "9876543210"
VENDOR_GST = "07ABCDE1234F1Z5"

INVOICE_FOLDER = "invoices"
INVOICE_FILE = "invoice_history.xlsx"
CUSTOMER_FILE = "customers.xlsx"

items = []
os.makedirs(INVOICE_FOLDER, exist_ok=True)

# ---------------- HELPERS ----------------
def invoice_no():
    return datetime.now().strftime("INV%Y%m%d%H%M%S")

def save_customer(cname, cname_company, addr, phone):
    if not os.path.exists(CUSTOMER_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Company", "Name", "Address", "Phone"])
        wb.save(CUSTOMER_FILE)

    wb = load_workbook(CUSTOMER_FILE)
    ws = wb.active
    ws.append([cname_company, cname, addr, phone])
    wb.save(CUSTOMER_FILE)

def save_invoice(inv, customer, total):
    if not os.path.exists(INVOICE_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Invoice No", "Customer", "Date", "Total"])
        wb.save(INVOICE_FILE)

    wb = load_workbook(INVOICE_FILE)
    ws = wb.active
    ws.append([inv, customer, datetime.now().strftime("%d-%m-%Y"), total])
    wb.save(INVOICE_FILE)

# ---------------- ADD ITEM ----------------
def add_item():
    name = entry_item.get()
    qty = entry_qty.get()
    price = entry_price.get()

    if not name or not qty or not price:
        messagebox.showerror("Error", "Item details missing")
        return

    total = int(qty) * float(price)
    items.append((name, qty, price, total))
    listbox.insert(tk.END, f"{name} | {qty} x {price} = {total}")

    entry_item.delete(0, tk.END)
    entry_qty.delete(0, tk.END)
    entry_price.delete(0, tk.END)

# ---------------- PDF ----------------
def generate_invoice():
    if not items:
        messagebox.showerror("Error", "Add at least one item")
        return

    cust_company = entry_cust_company.get()
    cust_name = entry_cust_name.get()
    cust_addr = entry_cust_addr.get()
    cust_phone = entry_cust_phone.get()

    inv = invoice_no()
    path = f"{INVOICE_FOLDER}/{inv}.pdf"
    c = canvas.Canvas(path, pagesize=A4)

    # LOGO
    if os.path.exists("company_logo.png"):
        c.drawImage("company_logo.png", 1.5*cm, 27*cm, 3*cm, 3*cm)

    # VENDOR
    c.setFont("Helvetica-Bold", 14)
    c.drawString(6*cm, 28*cm, VENDOR_NAME)
    c.setFont("Helvetica", 9)
    c.drawString(6*cm, 27.4*cm, VENDOR_ADDR)
    c.drawString(6*cm, 26.9*cm, f"Phone: {VENDOR_PHONE}")
    c.drawString(6*cm, 26.4*cm, f"GST: {VENDOR_GST}")

    # CUSTOMER
    c.setFont("Helvetica-Bold", 10)
    c.drawString(1.5*cm, 25.5*cm, "Bill To:")
    c.setFont("Helvetica", 9)
    c.drawString(1.5*cm, 25*cm, cust_company)
    c.drawString(1.5*cm, 24.5*cm, cust_name)
    c.drawString(1.5*cm, 24*cm, cust_addr)
    c.drawString(1.5*cm, 23.5*cm, cust_phone)

    c.drawString(14*cm, 25.5*cm, f"Invoice: {inv}")
    c.drawString(14*cm, 25*cm, f"Date: {datetime.now().strftime('%d-%m-%Y')}")

    # TABLE
    y = 22*cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(1.5*cm, y, "Item")
    c.drawString(9*cm, y, "Qty")
    c.drawString(11*cm, y, "Rate")
    c.drawString(14*cm, y, "Total")

    c.setFont("Helvetica", 10)
    subtotal = 0
    for i in items:
        y -= 0.7*cm
        c.drawString(1.5*cm, y, i[0])
        c.drawString(9*cm, y, str(i[1]))
        c.drawString(11*cm, y, str(i[2]))
        c.drawString(14*cm, y, str(i[3]))
        subtotal += i[3]

    cgst = subtotal * 0.09
    sgst = subtotal * 0.09
    grand = subtotal + cgst + sgst

    y -= 1.2*cm
    c.drawString(11*cm, y, "CGST 9%")
    c.drawString(14*cm, y, f"{cgst:.2f}")

    y -= 0.7*cm
    c.drawString(11*cm, y, "SGST 9%")
    c.drawString(14*cm, y, f"{sgst:.2f}")

    y -= 0.8*cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(11*cm, y, "Grand Total")
    c.drawString(14*cm, y, f"{grand:.2f}")

    c.save()

    save_customer(cust_name, cust_company, cust_addr, cust_phone)
    save_invoice(inv, cust_name, grand)

    items.clear()
    listbox.delete(0, tk.END)
    messagebox.showinfo("Success", f"Invoice Created\n{path}")

# ---------------- GUI ----------------
root = tk.Tk()
root.title("Invoice Generator")
root.geometry("480x650")

tk.Label(root, text="Invoice Generator", font=("Arial", 16, "bold")).pack(pady=10)

tk.Label(root, text="Customer Company (optional)").pack()
entry_cust_company = tk.Entry(root, width=50)
entry_cust_company.pack()

tk.Label(root, text="Customer Name").pack()
entry_cust_name = tk.Entry(root, width=50)
entry_cust_name.pack()

tk.Label(root, text="Customer Address").pack()
entry_cust_addr = tk.Entry(root, width=50)
entry_cust_addr.pack()

tk.Label(root, text="Customer Phone").pack()
entry_cust_phone = tk.Entry(root, width=50)
entry_cust_phone.pack()

tk.Label(root, text="Item Name").pack(pady=5)
entry_item = tk.Entry(root, width=40)
entry_item.pack()

entry_qty = tk.Entry(root, width=10)
entry_qty.pack()
entry_qty.insert(0, "Qty")

entry_price = tk.Entry(root, width=10)
entry_price.pack()
entry_price.insert(0, "Price")

tk.Button(root, text="Add Item", command=add_item).pack(pady=5)

listbox = tk.Listbox(root, width=60)
listbox.pack(pady=10)

tk.Button(root, text="Generate Invoice PDF", bg="green", fg="white",
          command=generate_invoice).pack(pady=15)

root.mainloop()
